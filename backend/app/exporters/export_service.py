"""
Export leads to CSV, Excel, JSON.
"""
import csv
import json
import io
from typing import List
from app.models.lead import Lead

EXPORT_FIELDS = [
    "id", "business_name", "category", "website", "email", "phone",
    "city", "country", "address", "score", "priority",
    "has_website", "https_enabled", "has_contact_page",
    "facebook", "instagram", "linkedin", "twitter",
    "score_explanation", "outreach_strategy", "source", "created_at",
]


def leads_to_csv(leads: List[Lead]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_FIELDS)
    writer.writeheader()
    for lead in leads:
        row = {f: getattr(lead, f, "") for f in EXPORT_FIELDS}
        row["created_at"] = str(row["created_at"])
        writer.writerow(row)
    return output.getvalue()


def leads_to_json(leads: List[Lead]) -> str:
    data = []
    for lead in leads:
        row = {f: getattr(lead, f, None) for f in EXPORT_FIELDS}
        row["created_at"] = str(row["created_at"])
        data.append(row)
    return json.dumps(data, indent=2, ensure_ascii=False)


def leads_to_excel(leads: List[Lead]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LeadRadar Leads"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    ws.append(EXPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for lead in leads:
        row = []
        for f in EXPORT_FIELDS:
            val = getattr(lead, f, "")
            if val is None:
                val = ""
            if hasattr(val, "isoformat"):
                val = str(val)
            row.append(val)
        ws.append(row)

    # Auto width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
